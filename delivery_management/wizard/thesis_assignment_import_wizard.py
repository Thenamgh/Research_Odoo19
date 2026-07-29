import base64
import re
from datetime import date, datetime
from io import BytesIO

from odoo import _, api, fields, models
from odoo.exceptions import UserError


class ThesisAssignmentImportWizard(models.TransientModel):
    _name = "thesis.assignment.import.wizard"
    _description = "Import danh sách phân công đồ án"

    REQUIRED_HEADERS = (
        "mã sinh viên",
        "họ và tên",
        "ngày sinh",
        "lớp",
        "tên đồ án",
        "giảng viên hướng dẫn",
        "mã giảng viên",
    )

    PROJECT_STATUS_RANK = {
        "draft": 0,
        "registered": 1,
        "assigned": 2,
        "accepted": 3,
        "in_progress": 4,
        "submitted": 5,
        "reviewing": 6,
        "defense_approved": 7,
        "defended": 8,
        "graduation_review": 9,
        "completed": 10,
        "cancelled": 11,
    }

    STUDENT_STATE_RANK = {
        "new": 0,
        "eligible": 1,
        "registered": 2,
        "assigned": 3,
        "submitted": 4,
        "defended": 5,
        "graduated": 6,
    }

    batch_id = fields.Many2one(
        comodel_name="thesis.batch",
        string="Đợt đồ án",
        required=True,
        ondelete="cascade",
        domain=[("state", "not in", ["done", "cancelled"])],
    )

    file_data = fields.Binary(
        string="File Excel",
        required=True,
        attachment=False,
    )

    file_name = fields.Char(
        string="Tên file",
    )

    validated = fields.Boolean(
        string="Đã kiểm tra",
        readonly=True,
    )

    row_count = fields.Integer(
        string="Số sinh viên",
        readonly=True,
    )

    validation_message = fields.Text(
        string="Kết quả kiểm tra",
        readonly=True,
    )

    @api.onchange("batch_id", "file_data", "file_name")
    def _onchange_import_input(self):
        """Khi đổi file hoặc đợt đồ án thì phải kiểm tra lại."""
        self.validated = False
        self.row_count = 0
        self.validation_message = False

    @staticmethod
    def _normalize_header(value):
        if value is None:
            return ""

        return re.sub(
            r"\s+",
            " ",
            str(value).strip().lower(),
        )

    @staticmethod
    def _normalize_text(value):
        if value is None:
            return ""

        if isinstance(value, float) and value.is_integer():
            value = int(value)

        return re.sub(
            r"\s+",
            " ",
            str(value).strip(),
        )

    @staticmethod
    def _same_text(first_value, second_value):
        first = ThesisAssignmentImportWizard._normalize_text(first_value).casefold()

        second = ThesisAssignmentImportWizard._normalize_text(second_value).casefold()

        return first == second

    @staticmethod
    def _parse_date(value):
        """Chuyển ngày Excel hoặc chuỗi thành đối tượng date."""
        if not value:
            return False

        if isinstance(value, datetime):
            return value.date()

        if isinstance(value, date):
            return value

        value = str(value).strip()

        accepted_formats = (
            "%d/%m/%Y",
            "%d-%m-%Y",
            "%Y-%m-%d",
        )

        for date_format in accepted_formats:
            try:
                return datetime.strptime(
                    value,
                    date_format,
                ).date()
            except ValueError:
                continue

        raise ValueError(
            "Ngày sinh phải có định dạng DD/MM/YYYY, " "DD-MM-YYYY hoặc YYYY-MM-DD."
        )

    def _raise_validation_errors(self, errors):
        """Hiển thị tối đa 50 lỗi để thông báo không quá dài."""
        if not errors:
            return

        displayed_errors = errors[:50]

        message = _(
            "File Excel chưa hợp lệ. Hệ thống chưa ghi dữ liệu.\n\n%s"
        ) % "\n".join("- %s" % error for error in displayed_errors)

        if len(errors) > 50:
            message += _("\n\nCòn %s lỗi khác chưa hiển thị.") % (len(errors) - 50)

        raise UserError(message)

    def _read_excel_rows(self):
        """Đọc Excel và chuẩn hóa dữ liệu từng dòng."""
        self.ensure_one()

        if not self.file_data:
            raise UserError(_("Vui lòng chọn file Excel."))

        if not self.file_name:
            raise UserError(_("Không xác định được tên file."))

        if not self.file_name.lower().endswith(".xlsx"):
            raise UserError(_("Hệ thống chỉ chấp nhận file có đuôi .xlsx."))

        try:
            from openpyxl import load_workbook
        except ImportError as error:
            raise UserError(
                _(
                    "Máy chủ chưa cài openpyxl.\n"
                    "Hãy chạy: python -m pip install openpyxl"
                )
            ) from error

        try:
            file_content = base64.b64decode(self.file_data)

            workbook = load_workbook(
                filename=BytesIO(file_content),
                read_only=True,
                data_only=True,
            )
        except Exception as error:
            raise UserError(
                _(
                    "Không thể đọc file Excel. "
                    "File có thể bị hỏng hoặc không phải XLSX hợp lệ."
                )
            ) from error

        rows = []
        errors = []

        try:
            worksheet = workbook.active

            header_row = next(
                worksheet.iter_rows(
                    min_row=1,
                    max_row=1,
                    values_only=True,
                ),
                None,
            )

            if not header_row:
                raise UserError(_("File Excel không có dòng tiêu đề."))

            normalized_headers = [self._normalize_header(value) for value in header_row]

            missing_headers = [
                header
                for header in self.REQUIRED_HEADERS
                if header not in normalized_headers
            ]

            if missing_headers:
                raise UserError(
                    _("File Excel thiếu các cột:\n- %s") % "\n- ".join(missing_headers)
                )

            duplicated_headers = [
                header
                for header in self.REQUIRED_HEADERS
                if normalized_headers.count(header) > 1
            ]

            if duplicated_headers:
                raise UserError(
                    _("File Excel có cột bị lặp:\n- %s")
                    % "\n- ".join(duplicated_headers)
                )

            column_indexes = {
                header: normalized_headers.index(header)
                for header in self.REQUIRED_HEADERS
            }

            for row_number, excel_row in enumerate(
                worksheet.iter_rows(
                    min_row=2,
                    values_only=True,
                ),
                start=2,
            ):
                # Bỏ qua dòng hoàn toàn trống
                if not any(cell not in (None, "") for cell in excel_row):
                    continue

                def get_cell(header):
                    index = column_indexes[header]

                    if index >= len(excel_row):
                        return None

                    return excel_row[index]

                student_code = self._normalize_text(get_cell("mã sinh viên"))
                student_name = self._normalize_text(get_cell("họ và tên"))
                class_code = self._normalize_text(get_cell("lớp"))
                project_title = self._normalize_text(get_cell("tên đồ án"))
                lecturer_name = self._normalize_text(get_cell("giảng viên hướng dẫn"))
                lecturer_code = self._normalize_text(get_cell("mã giảng viên"))

                if not student_code:
                    errors.append(
                        "Dòng %s, cột Mã sinh viên: không được để trống." % row_number
                    )

                if not student_name:
                    errors.append(
                        "Dòng %s, cột Họ và tên: không được để trống." % row_number
                    )

                if not class_code:
                    errors.append("Dòng %s, cột Lớp: không được để trống." % row_number)

                if not project_title:
                    errors.append(
                        "Dòng %s, cột Tên đồ án: không được để trống." % row_number
                    )

                if lecturer_code and not lecturer_name:
                    errors.append(
                        "Dòng %s: có mã giảng viên nhưng thiếu "
                        "tên giảng viên." % row_number
                    )

                if lecturer_name and not lecturer_code:
                    errors.append(
                        "Dòng %s: có tên giảng viên nhưng thiếu "
                        "mã giảng viên." % row_number
                    )

                try:
                    date_of_birth = self._parse_date(get_cell("ngày sinh"))

                    if not date_of_birth:
                        errors.append(
                            "Dòng %s, cột Ngày sinh: "
                            "không được để trống." % row_number
                        )
                except ValueError as error:
                    date_of_birth = False
                    errors.append("Dòng %s, cột Ngày sinh: %s" % (row_number, error))

                rows.append(
                    {
                        "row_number": row_number,
                        "student_code": student_code,
                        "student_name": student_name,
                        "date_of_birth": date_of_birth,
                        "class_code": class_code,
                        "project_title": project_title,
                        "lecturer_name": lecturer_name,
                        "lecturer_code": lecturer_code,
                    }
                )

            if not rows:
                raise UserError(_("File không có dòng dữ liệu sinh viên."))

            return rows, errors, worksheet.title

        finally:
            workbook.close()

    def _find_records_by_code(
        self,
        model_name,
        field_name,
        code,
    ):
        """Tìm mã không phân biệt chữ hoa/chữ thường."""
        return self.env[model_name].search([(field_name, "=ilike", code)])

    def _prepare_import(self, rows):
        """Kiểm tra dữ liệu với database nhưng chưa ghi dữ liệu."""
        errors = []
        plans = []
        seen_student_codes = {}

        for row in rows:
            row_number = row["row_number"]
            normalized_code = row["student_code"].casefold()

            if normalized_code in seen_student_codes:
                errors.append(
                    "Dòng %s: mã sinh viên %s bị lặp với dòng %s."
                    % (
                        row_number,
                        row["student_code"],
                        seen_student_codes[normalized_code],
                    )
                )
                continue

            seen_student_codes[normalized_code] = row_number

            classes = self._find_records_by_code(
                "academic.class",
                "code",
                row["class_code"],
            )

            if not classes:
                errors.append(
                    "Dòng %s, cột Lớp: không tìm thấy mã lớp %s."
                    % (row_number, row["class_code"])
                )
                continue

            if len(classes) > 1:
                errors.append(
                    "Dòng %s: mã lớp %s đang bị trùng trong hệ thống."
                    % (row_number, row["class_code"])
                )
                continue

            academic_class = classes[0]

            students = self._find_records_by_code(
                "thesis.student",
                "student_code",
                row["student_code"],
            )

            if len(students) > 1:
                errors.append(
                    "Dòng %s: mã sinh viên %s đang bị trùng "
                    "trong hệ thống." % (row_number, row["student_code"])
                )
                continue

            student = students[:1]

            lecturer = self.env["thesis.lecturer"]

            if row["lecturer_code"]:
                lecturers = self._find_records_by_code(
                    "thesis.lecturer",
                    "lecturer_code",
                    row["lecturer_code"],
                )

                if not lecturers:
                    errors.append(
                        "Dòng %s: không tìm thấy giảng viên mã %s."
                        % (row_number, row["lecturer_code"])
                    )
                    continue

                if len(lecturers) > 1:
                    errors.append(
                        "Dòng %s: mã giảng viên %s đang bị trùng."
                        % (row_number, row["lecturer_code"])
                    )
                    continue

                lecturer = lecturers[0]

                if not lecturer.active:
                    errors.append(
                        "Dòng %s: giảng viên %s đã ngừng hoạt động."
                        % (row_number, lecturer.display_name)
                    )
                    continue

                if not self._same_text(
                    lecturer.name,
                    row["lecturer_name"],
                ):
                    errors.append(
                        "Dòng %s: mã giảng viên %s thuộc về '%s', "
                        "không khớp với tên '%s'."
                        % (
                            row_number,
                            row["lecturer_code"],
                            lecturer.name,
                            row["lecturer_name"],
                        )
                    )
                    continue

            project = self.env["thesis.project"]

            if student:
                projects = self.env["thesis.project"].search(
                    [
                        ("student_id", "=", student.id),
                        ("batch_id", "=", self.batch_id.id),
                    ]
                )

                if len(projects) > 1:
                    errors.append(
                        "Dòng %s: sinh viên %s có nhiều hồ sơ "
                        "trong cùng đợt đồ án." % (row_number, row["student_code"])
                    )
                    continue

                project = projects[:1]

                other_active_project = self.env["thesis.project"].search(
                    [
                        ("student_id", "=", student.id),
                        ("batch_id", "!=", self.batch_id.id),
                        (
                            "status",
                            "not in",
                            ["completed", "cancelled"],
                        ),
                    ],
                    limit=1,
                )

                if other_active_project:
                    errors.append(
                        "Dòng %s: sinh viên %s đang có hồ sơ "
                        "hoạt động ở đợt khác: %s."
                        % (
                            row_number,
                            row["student_code"],
                            other_active_project.display_name,
                        )
                    )
                    continue

            if project and project.status in (
                "completed",
                "cancelled",
            ):
                errors.append(
                    "Dòng %s: hồ sơ %s đã ở trạng thái %s, "
                    "không được import lại."
                    % (
                        row_number,
                        project.display_name,
                        project.status,
                    )
                )
                continue

            # Nếu Excel không ghi GVHD thì giữ GVHD đã có,
            # không tự động xóa dữ liệu cũ.
            target_lecturer = lecturer

            if not target_lecturer and project:
                target_lecturer = project.supervisor_id

            if not target_lecturer and student and student.supervisor_id:
                target_lecturer = student.supervisor_id

            if project:
                current_rank = self.PROJECT_STATUS_RANK.get(
                    project.status,
                    0,
                )

                # Sau khi sinh viên đã nhận đề tài thì không cho
                # thay đổi tên đề tài hoặc GVHD bằng Excel.
                if current_rank >= self.PROJECT_STATUS_RANK["accepted"]:
                    if not self._same_text(
                        project.title,
                        row["project_title"],
                    ):
                        errors.append(
                            "Dòng %s: hồ sơ đã qua bước nhận đề tài, "
                            "không được thay đổi tên đồ án." % row_number
                        )
                        continue

                    if (
                        lecturer
                        and project.supervisor_id
                        and lecturer != project.supervisor_id
                    ):
                        errors.append(
                            "Dòng %s: hồ sơ đã qua bước nhận đề tài, "
                            "không được đổi giảng viên hướng dẫn." % row_number
                        )
                        continue

            target_status = "assigned" if target_lecturer else "registered"

            plans.append(
                {
                    "row": row,
                    "student": student,
                    "class": academic_class,
                    "lecturer": target_lecturer,
                    "project": project,
                    "target_status": target_status,
                }
            )

        self._check_lecturer_capacity(plans, errors)

        return plans, errors

    def _check_lecturer_capacity(self, plans, errors):
        """Kiểm tra giới hạn sau khi toàn bộ file được áp dụng."""
        lecturer_ids = {plan["lecturer"].id for plan in plans if plan["lecturer"]}

        for lecturer in self.env["thesis.lecturer"].browse(lecturer_ids):
            final_student_ids = set(lecturer.student_ids.ids)
            new_student_codes = set()
            related_rows = []

            for plan in plans:
                student = plan["student"]
                target_lecturer = plan["lecturer"]

                if (
                    student
                    and student.supervisor_id == lecturer
                    and target_lecturer != lecturer
                ):
                    final_student_ids.discard(student.id)

                if target_lecturer == lecturer:
                    related_rows.append(str(plan["row"]["row_number"]))

                    if student:
                        final_student_ids.add(student.id)
                    else:
                        new_student_codes.add(plan["row"]["student_code"].casefold())

            final_count = len(final_student_ids) + len(new_student_codes)

            if final_count > lecturer.max_students:
                errors.append(
                    "Giảng viên %s sẽ hướng dẫn %s sinh viên, "
                    "vượt giới hạn %s. Các dòng liên quan: %s."
                    % (
                        lecturer.display_name,
                        final_count,
                        lecturer.max_students,
                        ", ".join(related_rows),
                    )
                )

    def _open_wizard_again(self):
        return {
            "type": "ir.actions.act_window",
            "name": _("Import danh sách phân công"),
            "res_model": self._name,
            "res_id": self.id,
            "view_mode": "form",
            "target": "new",
        }

    def action_validate_file(self):
        """Kiểm tra toàn bộ file nhưng chưa ghi database."""
        self.ensure_one()

        rows, errors, worksheet_title = self._read_excel_rows()

        plans, database_errors = self._prepare_import(rows)
        errors.extend(database_errors)

        self._raise_validation_errors(errors)

        assigned_count = sum(1 for plan in plans if plan["lecturer"])

        self.write(
            {
                "validated": True,
                "row_count": len(rows),
                "validation_message": _(
                    "File hợp lệ.\n"
                    "Sheet: %s\n"
                    "Tổng số dòng: %s\n"
                    "Có GVHD: %s\n"
                    "Chưa có GVHD: %s\n"
                    "Đợt đồ án: %s"
                )
                % (
                    worksheet_title,
                    len(rows),
                    assigned_count,
                    len(rows) - assigned_count,
                    self.batch_id.display_name,
                ),
            }
        )

        return self._open_wizard_again()

    def action_import_file(self):
        """Import sinh viên và hồ sơ theo nguyên tắc all-or-nothing."""
        self.ensure_one()

        if not self.validated:
            raise UserError(_("Thầy cần kiểm tra file trước khi import."))

        # Đọc và kiểm tra lại để tránh dữ liệu thay đổi
        # sau lần nhấn Kiểm tra file.
        rows, errors, worksheet_title = self._read_excel_rows()

        plans, database_errors = self._prepare_import(rows)
        errors.extend(database_errors)

        self._raise_validation_errors(errors)

        created_students = 0
        updated_students = 0
        created_projects = 0
        updated_projects = 0

        today = fields.Date.context_today(self)

        for plan in plans:
            row = plan["row"]
            student = plan["student"]
            lecturer = plan["lecturer"]
            project = plan["project"]
            target_status = plan["target_status"]

            student_values = {
                "name": row["student_name"],
                "student_code": row["student_code"],
                "date_of_birth": row["date_of_birth"],
                "class_id": plan["class"].id,
                "class_name": plan["class"].code,
                "has_thesis_wish": True,
            }

            if lecturer:
                student_values["supervisor_id"] = lecturer.id

            if not student:
                student_values.update(
                    {
                        "state": target_status,
                        "eligible_date": today,
                        "eligible_checked_by": self.env.user.id,
                        "wish_date": today,
                    }
                )

                student = self.env["thesis.student"].create(student_values)
                created_students += 1

            else:
                desired_student_state = target_status

                current_rank = self.STUDENT_STATE_RANK.get(
                    student.state,
                    0,
                )
                desired_rank = self.STUDENT_STATE_RANK.get(
                    desired_student_state,
                    0,
                )

                if desired_rank > current_rank:
                    student_values["state"] = desired_student_state

                if not student.has_thesis_wish:
                    student_values["wish_date"] = today

                if student.state == "new":
                    student_values.update(
                        {
                            "eligible_date": today,
                            "eligible_checked_by": (self.env.user.id),
                        }
                    )

                student.write(student_values)
                updated_students += 1

            project_values = {
                "title": row["project_title"],
                "batch_id": self.batch_id.id,
                "student_id": student.id,
            }

            if lecturer:
                project_values["supervisor_id"] = lecturer.id

            if not project:
                project_values["status"] = target_status

                if target_status == "assigned":
                    project_values["assigned_by"] = self.env.user.id

                self.env["thesis.project"].create(project_values)
                created_projects += 1

            else:
                current_rank = self.PROJECT_STATUS_RANK.get(
                    project.status,
                    0,
                )
                desired_rank = self.PROJECT_STATUS_RANK.get(
                    target_status,
                    0,
                )

                # Không hạ trạng thái hồ sơ đã tiến xa hơn.
                if desired_rank > current_rank:
                    project_values["status"] = target_status

                if target_status == "assigned" and not project.assigned_by:
                    project_values["assigned_by"] = self.env.user.id

                project.write(project_values)
                updated_projects += 1

        self.write(
            {
                "validated": False,
                "validation_message": _("Import thành công từ sheet %s.")
                % worksheet_title,
            }
        )

        return {
            "type": "ir.actions.client",
            "tag": "display_notification",
            "params": {
                "title": _("Import thành công"),
                "message": _(
                    "Tạo mới %s sinh viên; cập nhật %s sinh viên; "
                    "tạo %s hồ sơ; cập nhật %s hồ sơ."
                )
                % (
                    created_students,
                    updated_students,
                    created_projects,
                    updated_projects,
                ),
                "type": "success",
                "sticky": True,
                "next": {
                    "type": "ir.actions.act_window",
                    "name": _("Hồ sơ vừa import"),
                    "res_model": "thesis.project",
                    "view_mode": "list,form",
                    "domain": [("batch_id", "=", self.batch_id.id)],
                },
            },
        }
